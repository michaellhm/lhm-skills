#!/usr/bin/env python3
"""
Build a colour-coded sitemap spreadsheet for client review.

Usage:
    python3 build_sitemap_sheet.py sitemap_data.json output.xlsx

Input JSON shape — see sitemap_data.example.json in this folder.

{
  "client":  "Client Name",
  "version": "v5",
  "date":    "2026-07-21",
  "live_url_total": 218,
  "sections": [
    {"band": "CORE  (8 pages)",
     "rows": [
       {"level": 1, "page": "Home", "url": "/", "type": "Core",
        "keyword": "brand term", "vol": "—", "status": "Planned"}
     ]}
  ],
  "programmatic": {
    "label": "Area Pages",
    "group_field": "Serving Clinic",
    "rows": [{"name": "Maribyrnong", "url": "/locations/maribyrnong/",
              "group": "Ascot Vale", "type": "Area page",
              "status": "Migrate as-is"}]
  },
  "decisions": [
    {"item": "Staff CPT", "urls": "15", "blocks": "RESOLVED",
     "issue": "...", "recommendation": "..."}
  ],
  "counts": [["Editorial pages", "57"], ["Staff CPT pages", "15"]]
}

Levels: 1 = top-level nav / section hub, 2 = child, 3 = grandchild,
        0 = outstanding / not yet placed in the architecture.
Type "Utility" renders grey (footer-only pages).

Status strings are matched for emphasis:
  "NEW..."                  -> amber highlight
  "DECISION"/"REVIEW"/"collision"/"CONFIRM" -> red bold
  "RESOLVED"                -> green bold (Decisions tab)

Requires: openpyxl
After generating, run the xlsx skill's recalc.py to populate formula caches.
"""

import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"

# Palette. L2/L3 are deliberately very pale — a saturated mid-tone repeated
# across dozens of rows reads as noise and reviewers ask for it to be toned down.
BAND = PatternFill("solid", fgColor="12333F")
L1F = PatternFill("solid", fgColor="1F4E5F")
L2F = PatternFill("solid", fgColor="E1EDF2")
L3F = PatternFill("solid", fgColor="F5FAFC")
UTILF = PatternFill("solid", fgColor="E8E8E8")
PENDF = PatternFill("solid", fgColor="F6C85F")
NEWF = PatternFill("solid", fgColor="FDE9C9")
RESF = PatternFill("solid", fgColor="DFF0DC")
GREYF = PatternFill("solid", fgColor="F2F2F2")

GROUP_FILLS = [
    PatternFill("solid", fgColor="D6E9D5"),
    PatternFill("solid", fgColor="D3E5EC"),
    PatternFill("solid", fgColor="EFE0EC"),
    PatternFill("solid", fgColor="FBE7D3"),
    PatternFill("solid", fgColor="E4DFF0"),
]

_thin = Side(style="thin", color="B7C9D0")
BORD = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

RED = "B3261E"
GREEN = "1E6B2F"
AMBER = "8A5A00"


def _header(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = BAND
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.alignment = Alignment(vertical="center", horizontal="left")
        c.border = BORD
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_sitemap_tab(wb, data):
    ws = wb.active
    ws.title = "Sitemap"
    head = ["Level", "Page", "URL", "Type", "Primary Keyword", "Vol/mo", "Status"]
    _header(ws, head)
    n = len(head)
    r = 2

    for section in data["sections"]:
        ws.cell(row=r, column=1, value=section["band"])
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n)
        for c in range(1, n + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = BAND
            cell.border = BORD
        anchor = ws.cell(row=r, column=1)
        anchor.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        anchor.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[r].height = 22
        r += 1

        for row in section["rows"]:
            lvl = row.get("level", 2)
            status = row.get("status", "")
            values = [
                {1: "L1", 2: "L2", 3: "L3", 0: "—"}.get(lvl, "—"),
                row.get("page", ""),
                row.get("url", ""),
                row.get("type", ""),
                row.get("keyword", ""),
                row.get("vol", ""),
                status,
            ]
            for c, v in enumerate(values, 1):
                ws.cell(row=r, column=c, value=v)

            if lvl == 0:
                fill, bold, white, indent = PENDF, True, False, 1
            elif row.get("type") == "Utility":
                fill, bold, white, indent = UTILF, False, False, 1
            elif lvl == 1:
                fill, bold, white, indent = L1F, True, True, 0
            elif lvl == 2:
                fill, bold, white, indent = L2F, False, False, 2
            else:
                fill, bold, white, indent = L3F, False, False, 4

            for c in range(1, n + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill
                cell.font = Font(
                    name=FONT, size=10, bold=bold,
                    color="FFFFFF" if white else "1A1A1A",
                )
                cell.border = BORD
                cell.alignment = Alignment(
                    vertical="center", indent=indent if c == 2 else 0
                )

            if status.startswith("NEW"):
                sc = ws.cell(row=r, column=7)
                sc.fill = NEWF
                sc.font = Font(name=FONT, size=10, bold=True, color=AMBER)
            if any(k in status for k in ("DECISION", "REVIEW", "collision", "CONFIRM")):
                ws.cell(row=r, column=7).font = Font(
                    name=FONT, size=10, bold=True, color=RED
                )
            r += 1

    _widths(ws, [8, 42, 46, 14, 40, 12, 30])
    ws.auto_filter.ref = f"A1:{get_column_letter(n)}1"
    return ws


def build_programmatic_tab(wb, prog):
    if not prog or not prog.get("rows"):
        return None
    ws = wb.create_sheet(prog.get("label", "Programmatic"))
    head = ["Name", "URL", prog.get("group_field", "Group"), "Type", "Status"]
    _header(ws, head)

    groups = []
    for row in prog["rows"]:
        if row.get("group") not in groups:
            groups.append(row.get("group"))
    fills = {g: GROUP_FILLS[i % len(GROUP_FILLS)] for i, g in enumerate(groups)}

    r = 2
    for row in prog["rows"]:
        values = [
            row.get("name", ""), row.get("url", ""), row.get("group", ""),
            row.get("type", ""), row.get("status", ""),
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fills[row.get("group")]
            cell.font = Font(name=FONT, size=10, color="1A1A1A")
            cell.border = BORD
        r += 1

    total_row = r + 1
    ws.cell(row=total_row, column=1, value="Total").font = Font(
        name=FONT, size=10, bold=True
    )
    ws.cell(row=total_row, column=2, value=f"=COUNTA(B2:B{r - 1})").font = Font(
        name=FONT, size=10, bold=True
    )
    _widths(ws, [24, 44, 18, 26, 30])
    ws.auto_filter.ref = "A1:E1"
    return ws


def build_decisions_tab(wb, decisions):
    if not decisions:
        return None
    ws = wb.create_sheet("Decisions")
    head = ["#", "Item", "Live URLs", "Blocks sign-off?", "Issue", "Recommendation"]
    _header(ws, head)

    r = 2
    for i, d in enumerate(decisions, 1):
        blocks = d.get("blocks", "No")
        values = [
            i, d.get("item", ""), d.get("urls", ""), blocks,
            d.get("issue", ""), d.get("recommendation", ""),
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name=FONT, size=10, color="1A1A1A")
            cell.border = BORD
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        if blocks == "RESOLVED":
            fill, colour = RESF, GREEN
        elif blocks.upper() in ("YES", "SCOPE TO CONFIRM"):
            fill, colour = PENDF, RED
        else:
            fill, colour = GREYF, "1A1A1A"

        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=4).font = Font(
            name=FONT, size=10, bold=True, color=colour
        )
        ws.row_dimensions[r].height = 58
        r += 1

    _widths(ws, [5, 24, 11, 16, 62, 62])
    return ws


def build_legend_tab(wb, data, prog):
    ws = wb.create_sheet("Legend")
    _widths(ws, [32, 78])

    def legend(r, label, fill, desc, white=False, bold=False):
        a = ws.cell(row=r, column=1, value=label)
        a.fill = fill
        a.font = Font(
            name=FONT, size=10, bold=bold or white,
            color="FFFFFF" if white else "1A1A1A",
        )
        a.border = BORD
        a.alignment = Alignment(vertical="center", indent=1)
        b = ws.cell(row=r, column=2, value=desc)
        b.font = Font(name=FONT, size=10, color="1A1A1A")
        b.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 20

    ws.cell(row=1, column=1, value="Navigation hierarchy").font = Font(
        name=FONT, size=12, bold=True, color="12333F"
    )
    legend(3, "Section band", BAND, "Top-level grouping", white=True, bold=True)
    legend(4, "L1 — Level 1", L1F,
           "Top-level nav item or section hub", white=True, bold=True)
    legend(5, "L2 — Level 2", L2F, "Child page, one click from a hub")
    legend(6, "L3 — Level 3", L3F, "Grandchild page")
    legend(7, "Utility", UTILF, "Footer-only page, not in primary nav")
    legend(8, "Outstanding", PENDF,
           "Live content type not yet placed — see Decisions tab")
    legend(9, "NEW", NEWF, "New page with no existing equity")

    r = 11
    if prog and prog.get("rows"):
        ws.cell(row=r, column=1,
                value=f"{prog.get('label', 'Programmatic')} tab — colour by "
                      f"{prog.get('group_field', 'group').lower()}").font = Font(
            name=FONT, size=12, bold=True, color="12333F"
        )
        r += 2
        groups = []
        for row in prog["rows"]:
            if row.get("group") not in groups:
                groups.append(row.get("group"))
        for i, g in enumerate(groups):
            count = sum(1 for x in prog["rows"] if x.get("group") == g)
            legend(r, g, GROUP_FILLS[i % len(GROUP_FILLS)],
                   f"{count} pages grouped under {g}")
            r += 1
        r += 1

    ws.cell(row=r, column=1, value="Counts").font = Font(
        name=FONT, size=12, bold=True, color="12333F"
    )
    r += 1
    for label, value in data.get("counts", []):
        a = ws.cell(row=r, column=1, value=label)
        a.font = Font(name=FONT, size=10, bold=True, color="1A1A1A")
        a.alignment = Alignment(indent=1)
        ws.cell(row=r, column=2, value=value).font = Font(
            name=FONT, size=10, color="1A1A1A"
        )
        r += 1

    if data.get("live_url_total"):
        a = ws.cell(row=r, column=1, value="Live URLs on current site")
        a.font = Font(name=FONT, size=10, bold=True, color="1A1A1A")
        a.alignment = Alignment(indent=1)
        ws.cell(row=r, column=2, value=data["live_url_total"]).font = Font(
            name=FONT, size=10, color="1A1A1A"
        )
        r += 1

    r += 1
    note = ws.cell(
        row=r, column=1,
        value=f"{data.get('client', '')} — sitemap {data.get('version', '')}, "
              f"{data.get('date', '')}.",
    )
    note.font = Font(name=FONT, size=9, italic=True, color="666666")
    if data.get("source_note"):
        ws.cell(row=r + 1, column=1, value=data["source_note"]).font = Font(
            name=FONT, size=9, italic=True, color="666666"
        )
    return ws


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    with open(sys.argv[1]) as f:
        data = json.load(f)

    wb = Workbook()
    prog = data.get("programmatic")
    build_sitemap_tab(wb, data)
    build_programmatic_tab(wb, prog)
    build_decisions_tab(wb, data.get("decisions"))
    build_legend_tab(wb, data, prog)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    wb.save(sys.argv[2])
    print(f"Saved {sys.argv[2]}")
    print("Next: run the xlsx skill's recalc.py to populate formula caches.")


if __name__ == "__main__":
    main()
