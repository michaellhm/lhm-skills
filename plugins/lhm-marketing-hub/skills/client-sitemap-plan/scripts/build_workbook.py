#!/usr/bin/env python3
"""
Build the companion .xlsx workbook from the same sitemap.json used by
build_sitemap_html.py. Optional deliverable. The HTML page is the hero, but some
clients like a spreadsheet they can sort and annotate.

Usage:
    python build_workbook.py sitemap.json output.xlsx

Produces two tabs:
    "Proposed Sitemap" — every page with section, slug, status, searches/mo
    "Keyword Demand"   — the pages that carry a search volume, sorted high to low

Reads only the page tree, so it works for both prospect and client specs. It never
touches the opportunity block and never writes lead estimates into the workbook.

Requires openpyxl. This script writes values only, so no recalculation is needed.

NOTE: this file is duplicated in the client-sitemap-plan skill. The two copies are
identical by design. If you change one, change both.
"""
import json, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GREEN="2E6B3E"; LGREEN="E4EFE7"; WHITE="FFFFFF"; DARK="1F2D23"
HAVE="2E7D32"; ENH="9A6E00"; NEW="C62828"
thin=Side(style="thin", color="CFD8CF"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
STATUS_WORD={"have":"Existing","enh":"Enhance","new":"New"}
STATUS_COLOUR={"have":HAVE,"enh":ENH,"new":NEW}

def header(ws,row,n):
    for c in range(1,n+1):
        cell=ws.cell(row=row,column=c)
        cell.fill=PatternFill("solid",fgColor=GREEN)
        cell.font=Font(name="Arial",bold=True,color=WHITE,size=10)
        cell.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
        cell.border=border

def build(spec, out):
    brand=spec["brand"]; name=brand.get("name","Client")
    wb=openpyxl.Workbook()

    # ---- Proposed Sitemap ----
    ws=wb.active; ws.title="Proposed Sitemap"; ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:E1")
    ws["A1"]=f"Proposed Sitemap - {name}"
    ws["A1"].font=Font(name="Arial",bold=True,size=14,color=WHITE)
    ws["A1"].fill=PatternFill("solid",fgColor=GREEN)
    ws["A1"].alignment=Alignment(vertical="center"); ws.row_dimensions[1].height=24
    heads=["Section","Page","URL slug","Status","Searches/mo"]
    for j,h in enumerate(heads,1): ws.cell(row=2,column=j,value=h)
    header(ws,2,len(heads)); ws.row_dimensions[2].height=22
    r=3; kw=[]
    for sec in spec["sections"]:
        for card in sec["cards"]:
            for p in card["items"]:
                status=p.get("status","new"); vol=p.get("vol")
                vals=[sec["title"], p["name"], p.get("slug",""), STATUS_WORD.get(status,"New"),
                      (int(vol) if isinstance(vol,(int,float)) and vol else "")]
                for j,v in enumerate(vals,1):
                    cell=ws.cell(row=r,column=j,value=v); cell.border=border
                    cell.font=Font(name="Arial",size=9.5,color=DARK)
                    cell.alignment=Alignment(vertical="top",wrap_text=True)
                ws.cell(row=r,column=1).font=Font(name="Arial",size=9.5,bold=True,color=GREEN)
                ws.cell(row=r,column=4).font=Font(name="Arial",size=9.5,bold=True,color=STATUS_COLOUR.get(status,NEW))
                ws.cell(row=r,column=5).alignment=Alignment(horizontal="center",vertical="center")
                fill=LGREEN if r%2 else WHITE
                for j in range(1,6):
                    if not ws.cell(row=r,column=j).fill.fgColor.rgb or ws.cell(row=r,column=j).fill.fgColor.rgb=="00000000":
                        ws.cell(row=r,column=j).fill=PatternFill("solid",fgColor=fill)
                if isinstance(vol,(int,float)) and vol:
                    kw.append((p["name"], int(vol), STATUS_WORD.get(status,"New"), sec["title"]))
                r+=1
    for j,w in enumerate([26,34,40,12,13],1):
        ws.column_dimensions[get_column_letter(j)].width=w
    ws.freeze_panes="A3"; ws.auto_filter.ref=f"A2:E{r-1}"

    # ---- Keyword Demand ----
    ws2=wb.create_sheet("Keyword Demand"); ws2.sheet_view.showGridLines=False
    ws2.merge_cells("A1:D1")
    ws2["A1"]="Keyword Demand - pages carrying measured search volume"
    ws2["A1"].font=Font(name="Arial",bold=True,size=14,color=WHITE)
    ws2["A1"].fill=PatternFill("solid",fgColor=GREEN); ws2.row_dimensions[1].height=24
    heads2=["Page","Searches/mo","Status","Section"]
    for j,h in enumerate(heads2,1): ws2.cell(row=2,column=j,value=h)
    header(ws2,2,len(heads2))
    kw.sort(key=lambda x:-x[1])
    r=3
    for pagename,vol,status,sec in kw:
        for j,v in enumerate([pagename,vol,status,sec],1):
            cell=ws2.cell(row=r,column=j,value=v); cell.border=border
            cell.font=Font(name="Arial",size=9.5,color=DARK)
            cell.alignment=Alignment(vertical="top",wrap_text=True)
        ws2.cell(row=r,column=2).alignment=Alignment(horizontal="center")
        ws2.cell(row=r,column=2).font=Font(name="Arial",size=9.5,bold=True,color=GREEN)
        fill=LGREEN if r%2 else WHITE
        for j in range(1,5): ws2.cell(row=r,column=j).fill=PatternFill("solid",fgColor=fill)
        r+=1
    for j,w in enumerate([40,14,12,26],1):
        ws2.column_dimensions[get_column_letter(j)].width=w
    ws2.freeze_panes="A3"; ws2.auto_filter.ref=f"A2:D{r-1}"

    wb.save(out)
    print("Wrote", out, "-", len(kw), "keyworded pages")

def main():
    if len(sys.argv)<3:
        print("Usage: python build_workbook.py <sitemap.json> <output.xlsx>"); sys.exit(1)
    with open(sys.argv[1],encoding="utf-8") as f: spec=json.load(f)
    build(spec, sys.argv[2])

if __name__=="__main__":
    main()
